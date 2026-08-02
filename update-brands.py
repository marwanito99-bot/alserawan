#!/usr/bin/env python3
"""
MWM Brands Manager — update-brands.py
======================================
Run this script whenever you edit brands.xlsx:
    python update-brands.py

It will:
  1. Read brands.xlsx (creates it with sample data if it doesn't exist)
  2. Generate brands-data.json (read by the website)

USAGE:
    python update-brands.py             → update from Excel
    python update-brands.py --create    → recreate Excel from scratch (overwrites!)
"""

import json, sys, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "brands.xlsx")
JSON_PATH  = os.path.join(os.path.dirname(__file__), "mwm-equipments", "brands-data.json")

# ─────────────────────────────────────────────────────────────
#  INITIAL DATA — edit here then run --create, OR edit in Excel
# ─────────────────────────────────────────────────────────────
INITIAL_BRANDS = [
    # ─── TIER 1: OUR BRANDS ───────────────────────────────────
    {
        "brand_id": "makute",
        "name_en": "MAKUTE",
        "name_ar": "ماكوتي",
        "name_cn": "MAKUTE",
        "country_en": "China",
        "country_ar": "الصين",
        "flag": "🇨🇳",
        "tagline_en": "Professional Power Tools",
        "tagline_ar": "عدد كهربائية احترافية",
        "description_en": "MAKUTE is a professional power tools brand engineered for demanding trades and construction work. Their comprehensive range spans corded and cordless tools, outdoor equipment, fuel-powered machinery, hand tools, and accessories — delivering industrial-grade performance at competitive cost for MENA professionals.",
        "description_ar": "ماكوتي علامة تجارية متخصصة في العدد الكهربائية الاحترافية، مصممة للأعمال الصعبة والبناء. تشمل تشكيلتها الواسعة العدد السلكية واللاسلكية، معدات الهواء الطلق، المحركات، العدد اليدوية والملحقات — بأداء صناعي وبتكلفة تنافسية.",
        "logo_path": "images/brands/makute/logo.webp",
        "hero_path": "images/brands/makute/hero-banner.webp",
        "color": "#ccfd1b",
        "website": "https://www.makute.com",
        "featured": 1,
        "active": 1,
        "sort_order": 1,
    },
    {
        "brand_id": "mwm",
        "name_en": "MWM",
        "name_ar": "MWM",
        "name_cn": "MWM",
        "country_en": "China · Italy · Germany",
        "country_ar": "الصين · إيطاليا · ألمانيا",
        "flag": "🏭",
        "tagline_en": "Our Own Label — Premium Equipment",
        "tagline_ar": "علامتنا الخاصة — معدات متميزة",
        "description_en": "MWM is our private label brand, representing a curated selection of heavy equipment and industrial machinery sourced and quality-verified directly through our Yiwu operations. Each MWM product meets stringent standards set by 25+ years of sourcing expertise.",
        "description_ar": "MWM هي علامتنا التجارية الخاصة، تمثل مجموعة منتقاة من المعدات الثقيلة والآلات الصناعية التي يتم توريدها والتحقق من جودتها مباشرة عبر عملياتنا في يوو. كل منتج MWM يلتزم بمعايير صارمة بناءً على 25+ عاماً من خبرة التوريد.",
        "logo_path": "images/logo-full.png",
        "hero_path": "images/ag1150.jpg",
        "color": "#ccfd1b",
        "website": "https://www.alserawan.com",
        "featured": 1,
        "active": 1,
        "sort_order": 2,
    },
    # ─── TIER 2: INTERNATIONAL BRANDS WITH FULL ASSETS ────────
    {
        "brand_id": "sicam",
        "name_en": "SICAM",
        "name_ar": "سيكام",
        "name_cn": "SICAM",
        "country_en": "Italy",
        "country_ar": "إيطاليا",
        "flag": "🇮🇹",
        "tagline_en": "Tyre Workshop Equipment",
        "tagline_ar": "معدات ورشة الإطارات",
        "description_en": "SICAM is an Italian manufacturer of professional tyre workshop equipment with over 40 years of engineering experience. Their range covers tyre changers, wheel balancers, aligners, and vehicle lifts — built for high-volume service centres and fleet maintenance operations.",
        "description_ar": "سيكام شركة إيطالية متخصصة في تصنيع معدات ورش الإطارات الاحترافية، بخبرة هندسية تتجاوز 40 عاماً. تشمل منتجاتها ماكينات تغيير الإطارات، موازنة العجلات، محاذيات ورافعات المركبات.",
        "logo_path": "images/brands/sicam/logo-clean.png",
        "hero_path": "images/brands/sicam/hero-banner.jpg",
        "color": "#1d00f4",
        "website": "https://www.sicam.it",
        "featured": 1,
        "active": 1,
        "sort_order": 3,
    },
    {
        "brand_id": "ingco",
        "name_en": "INGCO",
        "name_ar": "إنجكو",
        "name_cn": "INGCO",
        "country_en": "China",
        "country_ar": "الصين",
        "flag": "🇨🇳",
        "tagline_en": "Tools for Every Professional",
        "tagline_ar": "عدد لكل محترف",
        "description_en": "INGCO is one of the world's fastest-growing professional tools brands, operating in over 60 countries. Their extensive range covers power tools, hand tools, power equipment, pneumatic tools, and accessories — all engineered to meet international quality standards at exceptional value.",
        "description_ar": "إنجكو واحدة من أسرع العلامات التجارية للعدد الاحترافية نمواً في العالم، حاضرة في أكثر من 60 دولة. تشمل تشكيلتها الواسعة العدد الكهربائية، اليدوية، معدات الطاقة، العدد الهوائية والملحقات — مطابقة للمعايير الدولية بقيمة استثنائية.",
        "logo_path": "images/brands/ingco/logo.png",
        "hero_path": "images/brands/ingco/hero-banner.jpg",
        "color": "#e31e24",
        "website": "https://www.ingco.com",
        "featured": 1,
        "active": 1,
        "sort_order": 4,
    },
    {
        "brand_id": "ronix",
        "name_en": "RONIX",
        "name_ar": "رونيكس",
        "name_cn": "RONIX",
        "country_en": "Germany · Iran",
        "country_ar": "ألمانيا · إيران",
        "flag": "🇩🇪",
        "tagline_en": "German Engineering, Global Reach",
        "tagline_ar": "هندسة ألمانية، انتشار عالمي",
        "description_en": "RONIX is a premium power tools brand engineered with German-standard precision and manufactured for global professional markets. Their catalogue spans over 2,000 SKUs across angle grinders, drills, saws, and automotive tools — trusted by contractors across 50+ countries.",
        "description_ar": "رونيكس علامة تجارية متميزة للعدد الكهربائية، مصممة بدقة ألمانية للأسواق الاحترافية العالمية. تضم كتالوجها أكثر من 2000 منتج تشمل الطواحين، المثاقب، المناشير وعدد السيارات — موثوق بها من قِبَل المقاولين في 50+ دولة.",
        "logo_path": "images/brands/ronix/logo.svg",
        "hero_path": "images/brands/ronix/hero-banner.webp",
        "color": "#e63e2f",
        "website": "https://www.ronixtools.com",
        "featured": 1,
        "active": 1,
        "sort_order": 5,
    },
    {
        "brand_id": "worx",
        "name_en": "WORX",
        "name_ar": "ووركس",
        "name_cn": "WORX",
        "country_en": "USA · China",
        "country_ar": "أمريكا · الصين",
        "flag": "🇺🇸",
        "tagline_en": "Power Share — One Battery, All Tools",
        "tagline_ar": "بطارية واحدة لجميع العدد",
        "description_en": "WORX is an innovative power tools brand pioneering the Power Share universal battery platform — one 20V battery powers over 75 tools. Their lineup combines ergonomic cordless tools, outdoor equipment, and robotic garden solutions designed for modern professionals and serious DIYers.",
        "description_ar": "ووركس علامة تجارية مبتكرة تقود منصة PowerShare للبطارية الشاملة — بطارية 20 فولت واحدة تشغل أكثر من 75 أداة. تجمع تشكيلتها العدد اللاسلكية، المعدات الخارجية وحلول الحدائق الآلية.",
        "logo_path": "images/brands/worx/logo.svg",
        "hero_path": "images/brands/worx/hero-banner.jpg",
        "color": "#ff6600",
        "website": "https://www.worx.com",
        "featured": 1,
        "active": 1,
        "sort_order": 6,
    },
    {
        "brand_id": "flex",
        "name_en": "FLEX",
        "name_ar": "فليكس",
        "name_cn": "FLEX",
        "country_en": "Germany",
        "country_ar": "ألمانيا",
        "flag": "🇩🇪",
        "tagline_en": "German-Engineered Professional Tools",
        "tagline_ar": "عدد احترافية بهندسة ألمانية",
        "description_en": "FLEX is a premium German manufacturer of professional power tools — specialising in angle grinders, polishers, and surface treatment equipment since 1922. Renowned for innovation and build quality, FLEX tools are trusted by metalworkers, automotive professionals, and construction tradespeople worldwide.",
        "description_ar": "فليكس شركة ألمانية متميزة في تصنيع العدد الكهربائية الاحترافية منذ عام 1922 — متخصصة في الطواحين الزاوية، الماكينات اللامعة ومعدات معالجة الأسطح. موثوق بها من قِبَل مختصي المعادن، مهنيي السيارات وعمال البناء حول العالم.",
        "logo_path": "",
        "hero_path": "images/brands/flex/hero-banner-new.jpg",
        "color": "#005baa",
        "website": "https://www.flex-tools.com",
        "featured": 1,
        "active": 1,
        "sort_order": 7,
    },
    {
        "brand_id": "metabo",
        "name_en": "METABO",
        "name_ar": "ميتابو",
        "name_cn": "METABO",
        "country_en": "Germany",
        "country_ar": "ألمانيا",
        "flag": "🇩🇪",
        "tagline_en": "Safety, Technology, Performance",
        "tagline_ar": "الأمان، التكنولوجيا، الأداء",
        "description_en": "Metabo is a world-leading German manufacturer of professional power tools and accessories with over 95 years of innovation. Known for their pioneering cordless technology, safety systems, and robust engineering — trusted by construction, metalworking, and woodworking professionals in over 120 countries.",
        "description_ar": "ميتابو شركة ألمانية رائدة في تصنيع العدد الكهربائية الاحترافية بأكثر من 95 عاماً من الابتكار. معروفة بتقنيات اللاسلكي الرائدة وأنظمة الأمان والهندسة المتينة — موثوق بها من المحترفين في أكثر من 120 دولة.",
        "logo_path": "images/brands/metabo/logo.png",
        "hero_path": "images/brands/metabo/hero-banner-resized.jpg",
        "color": "#224b44",
        "website": "https://www.metabo.com",
        "featured": 1,
        "active": 1,
        "sort_order": 8,
    },
    {
        "brand_id": "bosch",
        "name_en": "BOSCH",
        "name_ar": "بوش",
        "name_cn": "博世",
        "country_en": "Germany",
        "country_ar": "ألمانيا",
        "flag": "🇩🇪",
        "tagline_en": "Invented for Life",
        "tagline_ar": "مخترع من أجل الحياة",
        "description_en": "Bosch Professional is the world's leading manufacturer of power tools and accessories. Founded in Stuttgart in 1886, Bosch brings over 135 years of German engineering excellence to every tool — from 18V cordless systems and SDS rotary hammers to angle grinders, demolition tools, and the legendary L-BOXX storage system. Trusted by tradespeople in 160+ countries.",
        "description_ar": "بوش برو فيشنال هي الشركة الرائدة عالمياً في صناعة العدد الكهربائية والملحقات. تأسست في شتوتغارت عام 1886، وتجلب بوش أكثر من 135 عاماً من الهندسة الألمانية المتميزة لكل أداة — من منظومات 18V اللاسلكية ومطارق SDS الدوارة إلى طواحين الزاوية وعدد الهدم ونظام التخزين L-BOXX الأسطوري. موثوق بها من قِبل المحترفين في 160+ دولة.",
        "logo_path": "images/brands/bosch/logo.svg",
        "hero_path": "images/brands/bosch/hero-banner.jpg",
        "color": "#e20015",
        "website": "https://www.bosch-professional.com",
        "featured": 1,
        "active": 1,
        "sort_order": 9,
    },
    {
        "brand_id": "karcher",
        "name_en": "KÄRCHER",
        "name_ar": "كارشر",
        "name_cn": "卡赫",
        "country_en": "Germany",
        "country_ar": "ألمانيا",
        "flag": "🇩🇪",
        "tagline_en": "World No.1 in Cleaning Technology",
        "tagline_ar": "الأول عالمياً في تقنية التنظيف",
        "description_en": "Kärcher is the world's leading provider of cleaning technology, headquartered in Winnenden, Germany. With over 3,000 patents and products sold in 60+ countries, Kärcher delivers professional high-pressure cleaners, floor scrubbers, vacuum cleaners, steam cleaners, and industrial cleaning systems. The gold standard for professional cleaning across construction, automotive, food industry, and public sector.",
        "description_ar": "كارشر هي الشركة الرائدة عالمياً في تقنية التنظيف، ومقرها وينيندن، ألمانيا. مع أكثر من 3,000 براءة اختراع ومنتجات تُباع في أكثر من 60 دولة، توفر كارشر ضواغط المياه عالية الضغط الاحترافية وآلات تلميع الأرضيات ومكانس الكهرباء ومنظفات البخار وأنظمة التنظيف الصناعي.",
        "logo_path": "images/brands/karcher/logo.svg",
        "hero_path": "",
        "color": "#FFD700",
        "website": "https://www.kaercher.com/ae",
        "featured": 1,
        "active": 1,
        "sort_order": 10,
    },
    {
        "brand_id": "total",
        "name_en": "TOTAL",
        "name_ar": "توتال",
        "name_cn": "大有",
        "country_en": "China",
        "country_ar": "الصين",
        "flag": "🇨🇳",
        "tagline_en": "Power for Professionals",
        "tagline_ar": "قوة المحترفين",
        "description_en": "TOTAL is a fast-growing professional power tools brand with a comprehensive portfolio spanning over 1,000 SKUs across corded and cordless tools, pneumatic tools, hand tools, and measuring instruments. The P20S 20V and S12 12V battery platforms deliver consistent power across 200+ compatible tools.",
        "description_ar": "توتال علامة تجارية سريعة النمو في العدد الكهربائية الاحترافية بمحفظة شاملة تضم أكثر من 1,000 منتج في العدد السلكية واللاسلكية والهوائية واليدوية وأجهزة القياس. تقدم منصتا P20S 20V و S12 12V طاقة متسقة عبر أكثر من 200 أداة متوافقة.",
        "logo_path": "",
        "hero_path": "images/brands/total/hero-banner.jpg",
        "color": "#e5002b",
        "website": "https://www.totalbusiness.com",
        "featured": 0,
        "active": 1,
        "sort_order": 11,
    },
    {
        "brand_id": "crown",
        "name_en": "CROWN",
        "name_ar": "كراون",
        "name_cn": "皇冠",
        "country_en": "United Arab Emirates",
        "country_ar": "الإمارات العربية المتحدة",
        "flag": "🇦🇪",
        "tagline_en": "Tools That Work",
        "tagline_ar": "عدد تشتغل",
        "description_en": "CROWN is a UAE-based professional power tools brand with over 30 years of presence in the MENA market. With more than 500 product lines spanning angle grinders, drills, circular saws, jigsaws, planers, and accessories, CROWN tools are engineered for the trades professional. The B3 platform integrates 20V lithium battery technology across a wide cordless ecosystem.",
        "description_ar": "كراون علامة إماراتية في العدد الكهربائية الاحترافية بحضور يزيد على 30 عاماً في منطقة الشرق الأوسط وشمال أفريقيا. بأكثر من 500 خط إنتاج يشمل طواحين الزاوية والثقاقات والمناشير الدائرية والمناشير الإهليلجية والمسوّيات والملحقات.",
        "logo_path": "images/brands/crown/logo.webp",
        "hero_path": "images/brands/crown/hero-banner.png",
        "color": "#cc0000",
        "website": "https://crown-tools.com",
        "featured": 1,
        "active": 1,
        "sort_order": 12,
    },
    {
        "brand_id": "dongcheng",
        "name_en": "DONGCHENG",
        "name_ar": "دونغ شينغ",
        "name_cn": "东成",
        "country_en": "China",
        "country_ar": "الصين",
        "flag": "🇨🇳",
        "tagline_en": "East Cheng — Born to Build",
        "tagline_ar": "شرق الصين — خُلق للبناء",
        "description_en": "Dongcheng (DCK) is one of China's largest professional power tools manufacturers, founded in 1995 in Jiangsu Province. Exporting to 100+ countries, Dongcheng produces over 3,000 product models covering angle grinders, rotary hammers, circular saws, cordless systems, and heavy-duty industrial tools.",
        "description_ar": "دونغ شينغ (DCK) هي إحدى أكبر شركات تصنيع العدد الكهربائية الاحترافية في الصين، تأسست عام 1995 في مقاطعة جيانغسو. تصدر لأكثر من 100 دولة وتنتج أكثر من 3,000 موديل يشمل طواحين الزاوية ومطارق SDS والمناشير الدائرية والمنظومات اللاسلكية.",
        "logo_path": "images/brands/dongcheng/logo.png",
        "hero_path": "images/brands/dongcheng/hero-banner.jpg",
        "color": "#e84919",
        "website": "https://www.dongcheng.com",
        "featured": 0,
        "active": 1,
        "sort_order": 13,
    },
    {
        "brand_id": "emtop",
        "name_en": "EMTOP",
        "name_ar": "إمتوب",
        "name_cn": "EMTOP",
        "country_en": "China",
        "country_ar": "الصين",
        "flag": "🇨🇳",
        "tagline_en": "Professional Tools for Every Job",
        "tagline_ar": "عدد احترافية لكل عمل",
        "description_en": "EMTOP is a professional-grade power tools brand under the Total Tools Group, offering a complete ecosystem of 20V P20S and 12V S12 battery-powered cordless tools alongside corded tools, pneumatics, and hand tools. With 1,000+ SKUs designed for contractors and tradespeople.",
        "description_ar": "إمتوب هي علامة احترافية في العدد الكهربائية ضمن مجموعة توتال تولز، تقدم منظومة متكاملة من العدد اللاسلكية بمنصة 20V P20S و 12V S12 إلى جانب خط كامل من العدد السلكية والهوائية واليدوية.",
        "logo_path": "",
        "hero_path": "images/brands/emtop/hero-banner.png",
        "color": "#0057a8",
        "website": "https://www.emtop-tools.com",
        "featured": 0,
        "active": 1,
        "sort_order": 14,
    },
]

INITIAL_PRODUCTS = [
    # MAKUTE
    {"product_id":"mk-001","brand_id":"makute","name_en":"Power Tools — AC","name_ar":"عدد كهربائية سلكية","image_path":"images/brands/makute/power-tools-ac.jpg","category_en":"Power Tools","category_ar":"عدد كهربائية","is_new":0,"active":1},
    {"product_id":"mk-002","brand_id":"makute","name_en":"Cordless Tool Systems","name_ar":"منظومات لاسلكية","image_path":"images/brands/makute/cordless-tools.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"mk-003","brand_id":"makute","name_en":"Hand Tools","name_ar":"عدد يدوية","image_path":"images/brands/makute/hand-tools.jpg","category_en":"Hand Tools","category_ar":"عدد يدوية","is_new":0,"active":1},
    {"product_id":"mk-004","brand_id":"makute","name_en":"Outdoor Equipment","name_ar":"معدات هواء طلق","image_path":"images/brands/makute/outdoor-tools.jpg","category_en":"Outdoor","category_ar":"هواء طلق","is_new":0,"active":1},
    {"product_id":"mk-005","brand_id":"makute","name_en":"Fuel Equipment","name_ar":"معدات وقود","image_path":"images/brands/makute/fuel-equipment.jpg","category_en":"Fuel","category_ar":"وقود","is_new":0,"active":1},
    {"product_id":"mk-006","brand_id":"makute","name_en":"Accessories & Parts","name_ar":"ملحقات وقطع غيار","image_path":"images/brands/makute/accessories.jpg","category_en":"Accessories","category_ar":"ملحقات","is_new":0,"active":1},
    # MWM
    {"product_id":"mw-001","brand_id":"mwm","name_en":"Angle Grinder AG1150","name_ar":"طاحونة AG1150","image_path":"images/ag1150.jpg","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"mw-002","brand_id":"mwm","name_en":"Chain Saw CW20","name_ar":"منشار جنزيري CW20","image_path":"images/cw20.jpg","category_en":"Saws","category_ar":"مناشير","is_new":0,"active":1},
    {"product_id":"mw-003","brand_id":"mwm","name_en":"Band Saw SB250","name_ar":"منشار شريطي SB250","image_path":"images/sb250.jpg","category_en":"Saws","category_ar":"مناشير","is_new":0,"active":1},
    {"product_id":"mw-004","brand_id":"mwm","name_en":"Screw Driver SD612","name_ar":"مفك SD612","image_path":"images/sd612.jpg","category_en":"Drills","category_ar":"مثاقب","is_new":0,"active":1},
    {"product_id":"mw-005","brand_id":"mwm","name_en":"Scroll Saw SW108","name_ar":"منشار SW108","image_path":"images/sw108.jpg","category_en":"Saws","category_ar":"مناشير","is_new":0,"active":1},
    {"product_id":"mw-006","brand_id":"mwm","name_en":"Water Pump WP3000","name_ar":"طرمبة مياه WP3000","image_path":"images/wp3000.jpg","category_en":"Pumps","category_ar":"طرمبات","is_new":0,"active":1},
    # SICAM
    {"product_id":"sicam-1","brand_id":"sicam","name_en":"Tyre Changers","name_ar":"مبدلات الإطارات","image_path":"images/brands/sicam/tyre-changer-hero.png","category_en":"Tyre Changers","category_ar":"مبدلات الإطارات","is_new":0,"active":1},
    {"product_id":"sicam-2","brand_id":"sicam","name_en":"Wheel Balancers","name_ar":"موازنات العجلات","image_path":"images/brands/sicam/wheel-balancer.png","category_en":"Wheel Balancers","category_ar":"موازنات العجلات","is_new":0,"active":1},
    {"product_id":"sicam-3","brand_id":"sicam","name_en":"Wheel Aligners","name_ar":"محاذيات العجلات","image_path":"images/brands/sicam/wheel-aligner.png","category_en":"Wheel Aligners","category_ar":"محاذيات العجلات","is_new":0,"active":1},
    {"product_id":"sicam-4","brand_id":"sicam","name_en":"Vehicle Lifts","name_ar":"روافع المركبات","image_path":"images/brands/sicam/lift-category.png","category_en":"Lifts","category_ar":"روافع","is_new":0,"active":1},
    # INGCO
    {"product_id":"ingco-1","brand_id":"ingco","name_en":"20V Cordless Platform","name_ar":"منظومة 20 فولت لاسلكية","image_path":"images/brands/ingco/cordless-20v.png","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"ingco-2","brand_id":"ingco","name_en":"Drills & Screwdrivers","name_ar":"مثاقب ومفكات","image_path":"images/brands/ingco/s12-platform.png","category_en":"Drills","category_ar":"مثاقب","is_new":0,"active":1},
    {"product_id":"ingco-3","brand_id":"ingco","name_en":"Angle Grinders","name_ar":"طواحين زاوية","image_path":"","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"ingco-4","brand_id":"ingco","name_en":"Hand Tools & Measuring","name_ar":"عدد يدوية وقياس","image_path":"","category_en":"Hand Tools","category_ar":"عدد يدوية","is_new":0,"active":1},
    # RONIX
    {"product_id":"ronix-1","brand_id":"ronix","name_en":"Angle Grinders","name_ar":"طواحين زاوية","image_path":"images/brands/ronix/cat-03.webp","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"ronix-2","brand_id":"ronix","name_en":"Drills & Impact Tools","name_ar":"مثاقب وعدد دق","image_path":"images/brands/ronix/cat-04.webp","category_en":"Drills","category_ar":"مثاقب","is_new":0,"active":1},
    {"product_id":"ronix-3","brand_id":"ronix","name_en":"Professional Tool Systems","name_ar":"منظومات عدد احترافية","image_path":"images/brands/ronix/about-03.webp","category_en":"Systems","category_ar":"منظومات","is_new":0,"active":1},
    {"product_id":"ronix-4","brand_id":"ronix","name_en":"Automotive & Air Tools","name_ar":"عدد سيارات وهوائية","image_path":"","category_en":"Automotive","category_ar":"عدد سيارات","is_new":0,"active":1},
    # WORX
    {"product_id":"worx-1","brand_id":"worx","name_en":"SwitchDriver Drill","name_ar":"مثقاب SwitchDriver","image_path":"images/brands/worx/drill-switchdriver.jpg","category_en":"Drills","category_ar":"مثاقب","is_new":0,"active":1},
    {"product_id":"worx-2","brand_id":"worx","name_en":"Cordless Circular Saw","name_ar":"منشار دائري لاسلكي","image_path":"images/brands/worx/circular-saw.jpg","category_en":"Saws","category_ar":"مناشير","is_new":0,"active":1},
    {"product_id":"worx-3","brand_id":"worx","name_en":"Cordless Leaf Blower","name_ar":"نافخ أوراق لاسلكي","image_path":"images/brands/worx/blower-leafjet.jpg","category_en":"Outdoor","category_ar":"معدات خارجية","is_new":0,"active":1},
    {"product_id":"worx-4","brand_id":"worx","name_en":"Compact Cordless Drill","name_ar":"مثقاب لاسلكي مدمج","image_path":"images/brands/worx/cordless-drill.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    # FLEX
    {"product_id":"flex-1","brand_id":"flex","name_en":"Angle Grinders","name_ar":"طواحين زاوية","image_path":"images/brands/flex/angle-grinder-sm.webp","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"flex-2","brand_id":"flex","name_en":"Large Angle Grinders","name_ar":"طواحين زاوية كبيرة","image_path":"images/brands/flex/large-grinder-sm.webp","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"flex-3","brand_id":"flex","name_en":"Cordless Polishers","name_ar":"ماكينات تلميع لاسلكية","image_path":"images/brands/flex/polisher-sm.webp","category_en":"Polishers","category_ar":"ماكينات تلميع","is_new":0,"active":1},
    {"product_id":"flex-4","brand_id":"flex","name_en":"SDS Hammer Drills","name_ar":"مطارق SDS","image_path":"images/brands/flex/hammer-drill-sm.webp","category_en":"SDS Drills","category_ar":"مطارق SDS","is_new":0,"active":1},
    # METABO
    {"product_id":"metabo-1","brand_id":"metabo","name_en":"Cordless Tool Systems","name_ar":"منظومات عدد لاسلكية","image_path":"images/brands/metabo/cordless-system.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"metabo-2","brand_id":"metabo","name_en":"Innovation Platform","name_ar":"منصة الابتكار","image_path":"images/brands/metabo/innovation.jpg","category_en":"Systems","category_ar":"منظومات","is_new":0,"active":1},
    {"product_id":"metabo-3","brand_id":"metabo","name_en":"Angle Grinders","name_ar":"طواحين زاوية","image_path":"","category_en":"Grinders","category_ar":"طواحين","is_new":0,"active":1},
    {"product_id":"metabo-4","brand_id":"metabo","name_en":"SDS Rotary Hammers","name_ar":"مطارق SDS دوارة","image_path":"","category_en":"SDS Hammers","category_ar":"مطارق SDS","is_new":0,"active":1},
    # BOSCH
    {"product_id":"bosch-1","brand_id":"bosch","name_en":"18V Cordless System","name_ar":"منظومة 18V اللاسلكية","image_path":"images/brands/bosch/hero-banner.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"bosch-2","brand_id":"bosch","name_en":"Metalworking Tools","name_ar":"عدد تشغيل المعادن","image_path":"images/brands/bosch/metalworking.jpg","category_en":"Metalworking","category_ar":"تشغيل المعادن","is_new":0,"active":1},
    {"product_id":"bosch-3","brand_id":"bosch","name_en":"L-BOXX System","name_ar":"نظام L-BOXX","image_path":"images/brands/bosch/lboxx-contractor.jpg","category_en":"Storage","category_ar":"تخزين","is_new":0,"active":1},
    # KÄRCHER (no product images yet)
    # TOTAL
    {"product_id":"total-1","brand_id":"total","name_en":"P20S 20V Platform","name_ar":"منصة P20S 20V","image_path":"images/brands/total/hero-banner.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"total-2","brand_id":"total","name_en":"S12 12V Platform","name_ar":"منصة S12 12V","image_path":"images/brands/total/s12-banner.jpg","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    # CROWN
    {"product_id":"crown-1","brand_id":"crown","name_en":"Professional Power Tools","name_ar":"عدد كهربائية احترافية","image_path":"images/brands/crown/hero-banner.png","category_en":"Power Tools","category_ar":"عدد كهربائية","is_new":0,"active":1},
    {"product_id":"crown-2","brand_id":"crown","name_en":"B3 Cordless Platform","name_ar":"منصة B3 اللاسلكية","image_path":"images/brands/crown/b3-platform.png","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    # DONGCHENG
    {"product_id":"dongcheng-1","brand_id":"dongcheng","name_en":"Professional Tool Range","name_ar":"تشكيلة العدد الاحترافية","image_path":"images/brands/dongcheng/hero-banner.jpg","category_en":"Power Tools","category_ar":"عدد كهربائية","is_new":0,"active":1},
    {"product_id":"dongcheng-2","brand_id":"dongcheng","name_en":"International Trade Show","name_ar":"المعرض التجاري الدولي","image_path":"images/brands/dongcheng/dubai-big5.jpg","category_en":"Exhibition","category_ar":"معارض","is_new":0,"active":1},
    # EMTOP
    {"product_id":"emtop-1","brand_id":"emtop","name_en":"Professional Tool Range","name_ar":"تشكيلة العدد الاحترافية","image_path":"images/brands/emtop/hero-banner.png","category_en":"Power Tools","category_ar":"عدد كهربائية","is_new":0,"active":1},
    {"product_id":"emtop-2","brand_id":"emtop","name_en":"P20S Cordless Platform","name_ar":"منصة P20S اللاسلكية","image_path":"images/brands/emtop/p20s-hero.png","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
    {"product_id":"emtop-3","brand_id":"emtop","name_en":"S12 Compact Platform","name_ar":"منصة S12 المدمجة","image_path":"images/brands/emtop/s12-hero.png","category_en":"Cordless","category_ar":"لاسلكية","is_new":0,"active":1},
]

INITIAL_NEW_ARRIVALS = []  # Add new arrival items here (same structure as products)

# ─────────────────────────────────────────────────────────────
#  EXCEL CREATION
# ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1d00f4")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
ROW_ALT_FILL  = PatternFill("solid", fgColor="F5F5FF")
THIN          = Border(
    left=Side(style="thin",color="CCCCCC"),
    right=Side(style="thin",color="CCCCCC"),
    top=Side(style="thin",color="CCCCCC"),
    bottom=Side(style="thin",color="CCCCCC"),
)

def style_header(ws, headers, col_widths):
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

def style_row(ws, row_idx, num_cols):
    fill = ROW_ALT_FILL if row_idx % 2 == 0 else None
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        if fill:
            c.fill = fill
        c.border = THIN
        c.alignment = Alignment(vertical="top", wrap_text=True)

def create_excel():
    wb = Workbook()

    # ── Sheet 1: Brands ──
    ws_b = wb.active
    ws_b.title = "Brands"
    ws_b.freeze_panes = "B2"

    brand_headers = [
        "brand_id","name_en","name_ar","name_cn","country_en","country_ar",
        "flag","tagline_en","tagline_ar","description_en","description_ar",
        "logo_path","hero_path","color","website","featured","active","sort_order"
    ]
    brand_widths = [14,12,14,12,12,12,6,22,22,45,45,30,30,10,30,8,7,8]
    style_header(ws_b, brand_headers, brand_widths)

    for ri, brand in enumerate(INITIAL_BRANDS, 2):
        for ci, key in enumerate(brand_headers, 1):
            c = ws_b.cell(row=ri, column=ci, value=brand.get(key, ""))
            c.alignment = Alignment(vertical="top", wrap_text=True)
        style_row(ws_b, ri, len(brand_headers))

    # ── Sheet 2: Products ──
    ws_p = wb.create_sheet("Products")
    ws_p.freeze_panes = "B2"

    prod_headers = [
        "product_id","brand_id","name_en","name_ar","image_path",
        "category_en","category_ar","is_new","active"
    ]
    prod_widths = [12,12,30,30,38,16,16,7,7]
    style_header(ws_p, prod_headers, prod_widths)

    for ri, prod in enumerate(INITIAL_PRODUCTS, 2):
        for ci, key in enumerate(prod_headers, 1):
            ws_p.cell(row=ri, column=ci, value=prod.get(key,""))
        style_row(ws_p, ri, len(prod_headers))

    # ── Sheet 3: New Arrivals ──
    ws_n = wb.create_sheet("New Arrivals")
    ws_n.freeze_panes = "B2"

    na_headers = [
        "product_id","brand_id","name_en","name_ar","image_path",
        "category_en","category_ar","active"
    ]
    na_widths = [12,12,30,30,38,16,16,7]
    style_header(ws_n, na_headers, na_widths)

    # Add usage note
    ws_n["A2"] = "Add new products here. Set active=1 to show them on the site."
    ws_n["A2"].font = Font(italic=True, color="888888")

    # ── Sheet 4: Instructions ──
    ws_i = wb.create_sheet("Instructions")
    instructions = [
        ("HOW TO USE THIS FILE", True),
        ("", False),
        ("1. Edit the Brands sheet to add/edit/remove brands.", False),
        ("   - Set active=0 to hide a brand without deleting it", False),
        ("   - Set featured=1 to give a brand a larger card", False),
        ("   - logo_path and hero_path: paths relative to site root (e.g. images/brands/makute/logo.webp)", False),
        ("   - color: hex color used for the brand accent (e.g. #ccfd1b)", False),
        ("", False),
        ("2. Edit the Products sheet to add/edit/remove products.", False),
        ("   - brand_id must match a brand_id in the Brands sheet", False),
        ("   - image_path: path relative to site root", False),
        ("   - Leave image_path empty for text-only product cards", False),
        ("", False),
        ("3. Edit the New Arrivals sheet to feature specific new products.", False),
        ("   - These appear in the 'New Arrivals' section at the top of the page", False),
        ("", False),
        ("4. AFTER EDITING: Save this file, then run:", False),
        ("       python update-brands.py", False),
        ("   This will update the website data automatically.", False),
        ("", False),
        ("IMAGE PATH FORMAT:", True),
        ("   images/brands/makute/logo.webp     ← for brand images", False),
        ("   images/ag1150.jpg                  ← for product images in root images folder", False),
    ]
    for row, (text, bold) in enumerate(instructions, 1):
        c = ws_i.cell(row=row, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12, color="1d00f4")
        else:
            c.font = Font(size=10)
    ws_i.column_dimensions["A"].width = 75

    wb.save(EXCEL_PATH)
    print(f"✅ Created Excel: {EXCEL_PATH}")

# ─────────────────────────────────────────────────────────────
#  READ EXCEL → GENERATE JSON
# ─────────────────────────────────────────────────────────────
def read_excel_and_generate_json():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ brands.xlsx not found at {EXCEL_PATH}")
        print("   Run: python update-brands.py --create  to create it")
        sys.exit(1)

    wb = load_workbook(EXCEL_PATH, data_only=True)

    # Read brands
    ws_b = wb["Brands"]
    brand_headers = [c.value for c in ws_b[1]]
    brands = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # skip empty rows
            continue
        brand = dict(zip(brand_headers, row))
        # normalize types
        brand["featured"] = int(brand.get("featured") or 0)
        brand["active"]   = int(brand.get("active")   or 0)
        brand["sort_order"] = int(brand.get("sort_order") or 99)
        brands.append(brand)

    brands.sort(key=lambda b: b.get("sort_order", 99))

    # Read products
    ws_p = wb["Products"]
    prod_headers = [c.value for c in ws_p[1]]
    products = []
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        prod = dict(zip(prod_headers, row))
        prod["is_new"] = int(prod.get("is_new") or 0)
        prod["active"] = int(prod.get("active") or 0)
        products.append(prod)

    # Read new arrivals
    ws_n = wb["New Arrivals"]
    na_headers = [c.value for c in ws_n[1]]
    new_arrivals = []
    for row in ws_n.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]).startswith("Add new"):
            continue
        item = dict(zip(na_headers, row))
        if int(item.get("active") or 0):
            new_arrivals.append(item)

    # Attach products to brands
    for brand in brands:
        bid = brand["brand_id"]
        brand["products"] = [
            p for p in products
            if p.get("brand_id") == bid and p.get("active") == 1
        ]

    output = {
        "brands": [b for b in brands if b.get("active") == 1],
        "new_arrivals": new_arrivals,
        "_generated": "Run update-brands.py to regenerate",
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    active_count  = len(output["brands"])
    product_count = sum(len(b["products"]) for b in output["brands"])
    print(f"✅ Generated: {JSON_PATH}")
    print(f"   Brands: {active_count}  |  Products: {product_count}  |  New Arrivals: {len(new_arrivals)}")

# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    force_create = "--create" in sys.argv

    if force_create or not os.path.exists(EXCEL_PATH):
        if force_create:
            print("🔄 Recreating Excel from scratch...")
        create_excel()

    read_excel_and_generate_json()
    print("\nDone! Push your changes to GitHub to publish.")
